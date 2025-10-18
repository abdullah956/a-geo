import logging
import math
from rest_framework import status, generics, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.pagination import PageNumberPagination
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db.models import Q, Count
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema, OpenApiExample
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import AttendanceSession, Attendance, AttendanceToken
from .token_utils import generate_token, verify_token, refresh_token, deactivate_session_tokens
from .serializers import (
    AttendanceSessionSerializer, AttendanceSessionCreateSerializer,
    AttendanceSerializer, AttendanceMarkSerializer,
    AttendanceSessionDetailSerializer, AttendanceSessionListSerializer,
    AttendanceStatsSerializer
)
from courses.models import Course, Enrollment

# Get logger instances
logger = logging.getLogger('attendance')
api_logger = logging.getLogger('api')


class SessionPagination(PageNumberPagination):
    """Custom pagination for attendance sessions"""
    page_size = 5
    page_size_query_param = 'page_size'
    max_page_size = 100


class IsTeacherOrAdmin(permissions.BasePermission):
    """
    Custom permission for teacher-specific attendance views
    """
    def has_permission(self, request, view):
        logger.info(f"Permission check for user: {request.user.email}, authenticated: {request.user.is_authenticated}, role: {request.user.role}")
        has_permission = request.user.is_authenticated and (
            request.user.is_superuser or
            request.user.role in ['admin', 'teacher']
        )
        logger.info(f"Permission result: {has_permission}")
        return has_permission


def calculate_distance(lat1, lon1, lat2, lon2):
    """
    Calculate distance between two coordinates using Haversine formula
    Returns distance in meters
    """
    R = 6371000  # Earth's radius in meters
    
    lat1_rad = math.radians(float(lat1))
    lat2_rad = math.radians(float(lat2))
    delta_lat = math.radians(float(lat2) - float(lat1))
    delta_lon = math.radians(float(lon2) - float(lon1))
    
    a = (math.sin(delta_lat / 2) ** 2 + 
         math.cos(lat1_rad) * math.cos(lat2_rad) * 
         math.sin(delta_lon / 2) ** 2)
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    
    return R * c


@extend_schema(
    operation_id='create_attendance_session',
    summary='Create attendance session',
    description='Create a new attendance session for a course. Only teachers can create sessions.',
    request=AttendanceSessionCreateSerializer,
    responses={
        201: AttendanceSessionSerializer,
        400: OpenApiExample(
            'Validation Error',
            value={'course_id': ['This field is required.']}
        ),
        403: OpenApiExample(
            'Forbidden',
            value={'detail': 'You do not have permission to perform this action.'}
        )
    }
)
class AttendanceSessionCreateView(generics.CreateAPIView):
    """
    Create a new attendance session - only teachers can create sessions
    """
    queryset = AttendanceSession.objects.all()
    serializer_class = AttendanceSessionCreateSerializer
    permission_classes = [IsTeacherOrAdmin]

    def perform_create(self, serializer):
        """Set the teacher field and validate course access"""
        course = serializer.validated_data.get('course')
        logger.info(f"Creating session for course: {course.code} - {course.title}, user: {self.request.user.email}")
        
        # Check if teacher is assigned to this course
        if self.request.user.role == 'teacher' and course.teacher != self.request.user:
            logger.error(f"User {self.request.user.email} not assigned to course {course.code}")
            raise PermissionDenied("You are not assigned to this course")
        
        serializer.save(teacher=self.request.user)
        logger.info(f"Attendance session created by {self.request.user.email} for course {course.code}")


@extend_schema(
    operation_id='list_attendance_sessions',
    summary='List attendance sessions',
    description='Get list of attendance sessions. Teachers see their sessions, students see sessions for their courses.',
    responses={
        200: AttendanceSessionListSerializer(many=True)
    }
)
class AttendanceSessionListView(generics.ListAPIView):
    """
    List attendance sessions based on user role
    """
    serializer_class = AttendanceSessionListSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = SessionPagination

    def get_queryset(self):
        """Filter sessions based on user role"""
        user = self.request.user
        
        if user.role == 'teacher':
            # Teachers see their own sessions
            return AttendanceSession.objects.filter(teacher=user).order_by('-started_at')
        elif user.role == 'student':
            # Students see sessions for their enrolled courses
            enrolled_courses = Course.objects.filter(
                enrollments__student=user,
                enrollments__is_active=True
            )
            return AttendanceSession.objects.filter(
                course__in=enrolled_courses
            ).order_by('-started_at')
        elif user.role == 'admin':
            # Admins see all sessions
            return AttendanceSession.objects.all().order_by('-started_at')
        
        return AttendanceSession.objects.none()


@extend_schema(
    operation_id='get_attendance_session',
    summary='Get attendance session details',
    description='Get detailed information about a specific attendance session.',
    responses={
        200: AttendanceSessionDetailSerializer,
        404: OpenApiExample(
            'Not Found',
            value={'detail': 'Not found.'}
        )
    }
)
class AttendanceSessionDetailView(generics.RetrieveAPIView):
    """
    Get detailed attendance session information
    """
    queryset = AttendanceSession.objects.all()
    serializer_class = AttendanceSessionDetailSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        """Check permissions for session access"""
        session = get_object_or_404(AttendanceSession, pk=self.kwargs['pk'])
        user = self.request.user
        
        # Check if user has access to this session
        if user.role == 'teacher' and session.teacher != user:
            raise PermissionDenied("You don't have access to this session")
        elif user.role == 'student':
            # Check if student is enrolled in the course
            if not Enrollment.objects.filter(
                student=user,
                course=session.course,
                is_active=True
            ).exists():
                raise PermissionDenied("You are not enrolled in this course")
        
        return session


@extend_schema(
    operation_id='end_attendance_session',
    summary='End attendance session',
    description='End an active attendance session. Only the teacher who started the session can end it.',
    responses={
        200: OpenApiExample(
            'Success',
            value={'message': 'Attendance session ended successfully'}
        ),
        403: OpenApiExample(
            'Forbidden',
            value={'detail': 'You do not have permission to perform this action.'}
        )
    }
)
@api_view(['POST'])
@permission_classes([IsTeacherOrAdmin])
def end_attendance_session(request, session_id):
    """
    End an active attendance session
    """
    session = get_object_or_404(AttendanceSession, id=session_id)
    
    # Check if teacher owns this session
    if request.user.role == 'teacher' and session.teacher != request.user:
        raise PermissionDenied("You can only end your own sessions")
    
    if session.status != 'active':
        return Response(
            {'error': 'Session is not active'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    session.end_session()
    logger.info(f"Attendance session {session.title} ended by {request.user.email}")
    
    return Response({
        'message': 'Attendance session ended successfully',
        'session': AttendanceSessionSerializer(session).data
    })


@extend_schema(
    operation_id='get_active_sessions',
    summary='Get active attendance sessions',
    description='Get all active attendance sessions for the current user.',
    responses={
        200: AttendanceSessionListSerializer(many=True)
    }
)
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_active_sessions(request):
    """
    Get active attendance sessions for the current user
    """
    user = request.user
    
    if user.role == 'student':
        # Get sessions for enrolled courses
        enrolled_courses = Course.objects.filter(
            enrollments__student=user,
            enrollments__is_active=True
        )
        sessions = AttendanceSession.objects.filter(
            course__in=enrolled_courses,
            status='active'
        ).order_by('-started_at')
    elif user.role == 'teacher':
        # Get teacher's active sessions
        sessions = AttendanceSession.objects.filter(
            teacher=user,
            status='active'
        ).order_by('-started_at')
    elif user.role == 'admin':
        # Get all active sessions
        sessions = AttendanceSession.objects.filter(
            status='active'
        ).order_by('-started_at')
    else:
        sessions = AttendanceSession.objects.none()
    
    serializer = AttendanceSessionListSerializer(sessions, many=True)
    return Response(serializer.data)


@extend_schema(
    operation_id='mark_attendance',
    summary='Mark attendance',
    description='Mark attendance for a student in an active session with location verification.',
    request=AttendanceMarkSerializer,
    responses={
        200: OpenApiExample(
            'Success',
            value={
                'message': 'Attendance marked successfully',
                'attendance': {
                    'id': 1,
                    'is_present': True,
                    'status': 'present',
                    'location_verified': True,
                    'distance_from_classroom': 25.5
                }
            }
        ),
        400: OpenApiExample(
            'Validation Error',
            value={'error': 'Location verification failed'}
        )
    }
)
@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def mark_attendance(request):
    """
    Mark attendance for a student with location verification
    """
    if request.user.role != 'student':
        return Response(
            {'error': 'Only students can mark attendance'}, 
            status=status.HTTP_403_FORBIDDEN
        )
    
    serializer = AttendanceMarkSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    session_id = serializer.validated_data['session_id']
    student_lat = serializer.validated_data['latitude']
    student_lon = serializer.validated_data['longitude']
    
    # Get the session
    session = get_object_or_404(AttendanceSession, id=session_id)
    
    # Check if session is active
    if not session.is_active:
        return Response(
            {'error': 'Attendance session is not active'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Check if student is enrolled in the course
    if not Enrollment.objects.filter(
        student=request.user,
        course=session.course,
        is_active=True
    ).exists():
        return Response(
            {'error': 'You are not enrolled in this course'}, 
            status=status.HTTP_403_FORBIDDEN
        )
    
    # Handle special case where coordinates are 0,0 (no location permission or outside radius)
    if student_lat == 0 and student_lon == 0:
        distance = float('inf')  # Set to infinity to indicate no location
        location_verified = False
    else:
        # Calculate distance from classroom
        distance = calculate_distance(
            session.classroom_latitude,
            session.classroom_longitude,
            student_lat,
            student_lon
        )
        
        # Check if location is within allowed radius
        location_verified = distance <= session.allowed_radius
    
    # Get or create attendance record
    attendance, created = Attendance.objects.get_or_create(
        session=session,
        student=request.user,
        defaults={
            'is_present': False,
            'status': 'absent'
        }
    )
    
    if not created and attendance.is_present:
        return Response(
            {'error': 'Attendance already marked'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Mark attendance
    attendance.mark_attendance(
        latitude=student_lat,
        longitude=student_lon,
        location_verified=location_verified,
        distance=distance
    )
    
    # Update status based on location verification
    if location_verified:
        attendance.status = 'present'
    else:
        attendance.status = 'absent'
        attendance.is_present = False
    
    attendance.save()
    
    logger.info(f"Attendance marked by {request.user.email} for session {session.title}")
    
    # Handle infinity distance in response
    response_distance = distance if distance != float('inf') else -1
    
    return Response({
        'message': 'Attendance marked successfully',
        'attendance': AttendanceSerializer(attendance).data,
        'location_verified': location_verified,
        'distance': response_distance,
        'allowed_radius': session.allowed_radius
    })


@extend_schema(
    operation_id='get_student_notifications',
    summary='Get student notifications',
    description='Get active sessions and notifications for students.',
    responses={
        200: OpenApiExample(
            'Student Notifications',
            value={
                'active_sessions': [
                    {
                        'id': 1,
                        'title': 'Lecture 1',
                        'course_code': 'CS101',
                        'classroom_name': 'Room 101',
                        'started_at': '2025-01-12T10:00:00Z',
                        'allowed_radius': 50
                    }
                ],
                'notifications': [
                    {
                        'type': 'attendance_session_started',
                        'message': 'New attendance session started for CS101',
                        'session_id': 1
                    }
                ]
            }
        )
    }
)
@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def get_student_notifications(request):
    """
    Get notifications and active sessions for students
    """
    if request.user.role != 'student':
        return Response(
            {'error': 'Only students can access notifications'}, 
            status=status.HTTP_403_FORBIDDEN
        )
    
    try:
        # Get active sessions for student's enrolled courses
        from courses.models import Enrollment
        enrolled_courses = Course.objects.filter(
            enrollments__student=request.user,
            enrollments__is_active=True
        )
        
        active_sessions = AttendanceSession.objects.filter(
            course__in=enrolled_courses,
            status='active'
        ).order_by('-started_at')
        
        # Check if student has already marked attendance for these sessions
        sessions_with_attendance = []
        for session in active_sessions:
            attendance_record = Attendance.objects.filter(
                session=session,
                student=request.user
            ).first()
            
            session_data = {
                'id': session.id,
                'title': session.title,
                'course_code': session.course.code,
                'classroom_name': session.classroom_name,
                'classroom_latitude': str(session.classroom_latitude),
                'classroom_longitude': str(session.classroom_longitude),
                'started_at': session.started_at.isoformat(),
                'allowed_radius': session.allowed_radius,
                'attendance_marked': attendance_record.is_present if attendance_record else False,
                'attendance_status': attendance_record.status if attendance_record else 'not_marked'
            }
            sessions_with_attendance.append(session_data)
        
        # Generate notifications for new sessions
        notifications = []
        for session in sessions_with_attendance:
            if not session['attendance_marked']:
                notifications.append({
                    'type': 'attendance_session_started',
                    'message': f"New attendance session started for {session['course_code']}",
                    'session_id': session['id'],
                    'title': session['title']
                })
        
        return Response({
            'active_sessions': sessions_with_attendance,
            'notifications': notifications,
            'total_sessions': len(sessions_with_attendance),
            'unmarked_sessions': len([s for s in sessions_with_attendance if not s['attendance_marked']])
        })
        
    except Exception as e:
        logger.error(f"Error getting student notifications: {str(e)}")
        return Response(
            {'error': 'Failed to get notifications'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@extend_schema(
    operation_id='get_attendance_stats',
    summary='Get attendance statistics',
    description='Get attendance statistics for teachers and admins.',
    responses={
        200: AttendanceStatsSerializer
    }
)
@api_view(['GET'])
@permission_classes([IsTeacherOrAdmin])
def get_attendance_stats(request):
    """
    Get attendance statistics
    """
    user = request.user
    
    if user.role == 'teacher':
        # Get teacher's sessions
        sessions = AttendanceSession.objects.filter(teacher=user)
    else:
        # Admin sees all sessions
        sessions = AttendanceSession.objects.all()
    
    total_sessions = sessions.count()
    active_sessions = sessions.filter(status='active').count()
    total_attendance = Attendance.objects.filter(session__in=sessions, is_present=True).count()
    
    # Calculate attendance rate
    total_possible_attendance = sessions.aggregate(
        total=Count('attendances')
    )['total'] or 0
    
    attendance_rate = (total_attendance / total_possible_attendance * 100) if total_possible_attendance > 0 else 0
    
    # Get recent sessions
    recent_sessions = sessions.order_by('-started_at')[:5]
    recent_sessions_data = AttendanceSessionListSerializer(recent_sessions, many=True).data
    
    stats = {
        'total_sessions': total_sessions,
        'active_sessions': active_sessions,
        'total_attendance_marked': total_attendance,
        'attendance_rate': round(attendance_rate, 2),
        'recent_sessions': recent_sessions_data
    }
    
    return Response(stats)


@extend_schema(
    operation_id='export_session_to_excel',
    summary='Export session attendance to Excel',
    description='Export detailed attendance data for a session to Excel file.',
    responses={
        200: OpenApiExample(
            'Success',
            value={'file': 'Excel file download'}
        ),
        403: OpenApiExample(
            'Forbidden',
            value={'detail': 'You do not have permission to perform this action.'}
        )
    }
)
@api_view(['GET'])
@permission_classes([IsTeacherOrAdmin])
def export_session_to_excel(request, session_id):
    """
    Export session attendance data to Excel
    """
    session = get_object_or_404(AttendanceSession, id=session_id)
    
    # Check if user has access to this session
    if request.user.role == 'teacher' and session.teacher != request.user:
        raise PermissionDenied("You don't have access to this session")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Add header information
    ws['A1'] = "Attendance Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Session: {session.title}"
    ws['A3'] = f"Course: {session.course.code} - {session.course.title}"
    ws['A4'] = f"Teacher: {session.teacher.get_full_name()}"
    ws['A5'] = f"Classroom: {session.classroom_name}"
    ws['A6'] = f"Started: {session.started_at.strftime('%Y-%m-%d %H:%M:%S')}"
    if session.ended_at:
        ws['A7'] = f"Ended: {session.ended_at.strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A8'] = f"Status: {session.status.upper()}"
    
    # Add column headers for attendance data
    headers = ['#', 'Student ID', 'Student Name', 'Email', 'Status', 'Marked At', 'Location Verified', 'Distance (m)']
    header_row = 10
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Get all enrolled students
    enrollments = Enrollment.objects.filter(
        course=session.course,
        is_active=True
    ).select_related('student')
    
    # Get attendance records
    attendance_records = Attendance.objects.filter(session=session).select_related('student')
    attendance_dict = {att.student.id: att for att in attendance_records}
    
    # Fill in student data
    row_num = header_row + 1
    for idx, enrollment in enumerate(enrollments, start=1):
        student = enrollment.student
        attendance = attendance_dict.get(student.id)
        
        ws.cell(row=row_num, column=1).value = idx
        ws.cell(row=row_num, column=2).value = student.username
        ws.cell(row=row_num, column=3).value = student.get_full_name()
        ws.cell(row=row_num, column=4).value = student.email
        
        if attendance and attendance.is_present:
            ws.cell(row=row_num, column=5).value = "PRESENT"
            ws.cell(row=row_num, column=5).font = Font(color="00B050", bold=True)
            ws.cell(row=row_num, column=6).value = attendance.marked_at.strftime('%Y-%m-%d %H:%M:%S') if attendance.marked_at else 'N/A'
            ws.cell(row=row_num, column=7).value = "Yes" if attendance.location_verified else "No"
            ws.cell(row=row_num, column=8).value = round(attendance.distance_from_classroom, 2) if attendance.distance_from_classroom else 'N/A'
        else:
            ws.cell(row=row_num, column=5).value = "ABSENT"
            ws.cell(row=row_num, column=5).font = Font(color="FF0000", bold=True)
            ws.cell(row=row_num, column=6).value = "Not Marked"
            ws.cell(row=row_num, column=7).value = "N/A"
            ws.cell(row=row_num, column=8).value = "N/A"
        
        row_num += 1
    
    # Add summary
    summary_row = row_num + 2
    ws.cell(row=summary_row, column=1).value = "Summary:"
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1).value = "Total Students:"
    ws.cell(row=summary_row + 1, column=2).value = enrollments.count()
    ws.cell(row=summary_row + 2, column=1).value = "Present:"
    ws.cell(row=summary_row + 2, column=2).value = session.attendance_count
    ws.cell(row=summary_row + 2, column=2).font = Font(color="00B050", bold=True)
    ws.cell(row=summary_row + 3, column=1).value = "Absent:"
    ws.cell(row=summary_row + 3, column=2).value = enrollments.count() - session.attendance_count
    ws.cell(row=summary_row + 3, column=2).font = Font(color="FF0000", bold=True)
    ws.cell(row=summary_row + 4, column=1).value = "Attendance Rate:"
    attendance_rate = (session.attendance_count / enrollments.count() * 100) if enrollments.count() > 0 else 0
    ws.cell(row=summary_row + 4, column=2).value = f"{attendance_rate:.1f}%"
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"attendance_{session.course.code}_{session.title}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    
    logger.info(f"Attendance data exported by {request.user.email} for session {session.id}")
    
    return response


# ============================================================================
# QR CODE TOKEN ENDPOINTS
# ============================================================================

@extend_schema(
    operation_id='generate_qr_token',
    summary='Generate QR code token for attendance session',
    description='Teachers can generate a QR code token for their active attendance session.',
    responses={
        200: OpenApiExample(
            'Success',
            value={
                'token': 'eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9...',
                'token_hash': 'abc123...',
                'expires_at': '2025-01-12T10:15:00Z',
                'qr_code': 'data:image/png;base64,...',
                'token_id': 1
            }
        ),
        403: OpenApiExample(
            'Forbidden',
            value={'detail': 'You do not have permission to perform this action.'}
        ),
        404: OpenApiExample(
            'Not Found',
            value={'detail': 'Not found.'}
        )
    }
)
@api_view(['POST'])
@permission_classes([IsTeacherOrAdmin])
def generate_qr_token_view(request, session_id):
    """
    Generate a QR code token for an attendance session
    """
    session = get_object_or_404(AttendanceSession, id=session_id)
    
    # Check if user is the teacher of this session
    if request.user.role == 'teacher' and session.teacher != request.user:
        raise PermissionDenied("You are not the teacher of this session")
    
    # Check if session is active
    if not session.is_active:
        return Response(
            {'error': 'Session is not active'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Get duration from request (default 10 minutes)
    duration_minutes = request.data.get('duration_minutes', 10)
    
    # Generate token
    token_data = generate_token(session, duration_minutes=duration_minutes)
    
    logger.info(f"QR token generated for session {session.id} by {request.user.email}")
    
    return Response(token_data, status=status.HTTP_200_OK)


@extend_schema(
    operation_id='refresh_qr_token',
    summary='Refresh QR code token',
    description='Teachers can refresh/regenerate a QR code token for their session.',
    responses={
        200: OpenApiExample(
            'Success',
            value={
                'token': 'eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9...',
                'token_hash': 'abc123...',
                'expires_at': '2025-01-12T10:15:00Z',
                'qr_code': 'data:image/png;base64,...',
                'token_id': 2
            }
        )
    }
)
@api_view(['POST'])
@permission_classes([IsTeacherOrAdmin])
def refresh_qr_token_view(request, session_id):
    """
    Refresh QR code token for a session
    """
    session = get_object_or_404(AttendanceSession, id=session_id)
    
    # Check if user is the teacher of this session
    if request.user.role == 'teacher' and session.teacher != request.user:
        raise PermissionDenied("You are not the teacher of this session")
    
    # Get old token from request (optional)
    old_token = request.data.get('old_token', None)
    
    # Refresh token
    token_data = refresh_token(session, old_token)
    
    logger.info(f"QR token refreshed for session {session.id} by {request.user.email}")
    
    return Response(token_data, status=status.HTTP_200_OK)


@extend_schema(
    operation_id='verify_qr_token',
    summary='Verify QR code token and mark attendance',
    description='Students can verify a QR code token and mark their attendance.',
    request={
        'application/json': {
            'type': 'object',
            'properties': {
                'token': {'type': 'string', 'description': 'JWT token from QR code'},
                'latitude': {'type': 'number', 'description': 'Student latitude (optional)'},
                'longitude': {'type': 'number', 'description': 'Student longitude (optional)'}
            },
            'required': ['token']
        }
    },
    responses={
        200: OpenApiExample(
            'Success',
            value={
                'message': 'Attendance marked successfully',
                'attendance': {...},
                'session': {...}
            }
        ),
        400: OpenApiExample(
            'Bad Request',
            value={'error': 'Invalid or expired token'}
        )
    }
)
@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def verify_qr_token_view(request):
    """
    Verify QR code token and mark attendance
    """
    if request.user.role != 'student':
        return Response(
            {'error': 'Only students can mark attendance'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    token_string = request.data.get('token')
    if not token_string:
        return Response(
            {'error': 'Token is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Verify token
    payload = verify_token(token_string)
    if not payload:
        return Response(
            {'error': 'Invalid or expired token'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Get session from payload
    session_id = payload.get('session_id')
    session = get_object_or_404(AttendanceSession, id=session_id)
    
    # Check if session is still active
    if not session.is_active:
        return Response(
            {'error': 'Session is no longer active'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Check if student is enrolled
    if not Enrollment.objects.filter(
        student=request.user,
        course=session.course,
        is_active=True
    ).exists():
        return Response(
            {'error': 'You are not enrolled in this course'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    # Get location data (optional)
    latitude = request.data.get('latitude', 0)
    longitude = request.data.get('longitude', 0)
    
    # Calculate distance if location provided
    distance = 0
    location_verified = True  # QR code implies presence
    
    if latitude != 0 and longitude != 0:
        from .views import calculate_distance
        distance = calculate_distance(
            session.classroom_latitude,
            session.classroom_longitude,
            latitude,
            longitude
        )
        location_verified = distance <= session.allowed_radius
    
    # Create or update attendance
    attendance, created = Attendance.objects.get_or_create(
        session=session,
        student=request.user,
        defaults={
            'is_present': True,
            'status': 'present',
            'location_verified': location_verified,
            'student_latitude': latitude if latitude != 0 else None,
            'student_longitude': longitude if longitude != 0 else None,
            'distance_from_classroom': distance if distance > 0 else None,
            'marked_at': timezone.now()
        }
    )
    
    if not created:
        # Update existing attendance
        attendance.is_present = True
        attendance.status = 'present'
        attendance.location_verified = location_verified
        attendance.marked_at = timezone.now()
        if latitude != 0:
            attendance.student_latitude = latitude
            attendance.student_longitude = longitude
        if distance > 0:
            attendance.distance_from_classroom = distance
        attendance.save()
    
    logger.info(f"Attendance marked via QR code for {request.user.email} in session {session.id}")
    
    return Response({
        'message': 'Attendance marked successfully',
        'attendance': AttendanceSerializer(attendance).data,
        'session': {
            'id': session.id,
            'title': session.title,
            'course_code': session.course.code
        }
    }, status=status.HTTP_200_OK)
